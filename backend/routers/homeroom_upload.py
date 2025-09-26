from __future__ import annotations

from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional, Any
from io import BytesIO
from datetime import datetime, date, timedelta
from urllib.parse import quote

from database import get_db
from models import Student, User, HomeroomAssignment
from routers.auth import get_current_user, role_required

from openpyxl import Workbook, load_workbook

router = APIRouter()

EXPECTED_HEADERS = [
    "학년", "반", "번호", "성명", "학생개인번호", "성별", "생년월일",
    "주소", "비고", "연락처", "보호자1연락처", "보호자2연락처", "username",
]

# ------------------------- 파서 유틸 -------------------------
def _safe_str(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s != "" else None

def _safe_int(v: Any) -> int | None:
    """
    엑셀에서 숫자가 float로 오거나 문자열로 와도 변환.
    """
    if v is None:
        return None
    # 이미 int/float인 경우
    if isinstance(v, (int,)):
        return v
    if isinstance(v, float):
        try:
            return int(v)
        except Exception:
            return None
    # 문자열
    try:
        s = str(v).strip()
        if s == "":
            return None
        # 1.2 처럼 올 수도 있어 제거
        if "." in s:
            s = s.split(".", 1)[0]
        return int(s)
    except Exception:
        return None

def _excel_date_to_date(v: Any) -> date | None:
    """
    엑셀 셀에 날짜가 '값'으로 들어온 경우:
    - datetime/date 객체면 date로
    - float/int(엑셀 시리얼)면 1899-12-30 기준 가정하여 변환
    """
    if v is None:
        return None
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, (int, float)):
        # openpyxl이 시리얼 날짜를 숫자로 줄 수 있음: 1899-12-30 기준
        try:
            base = date(1899, 12, 30)
            return base + timedelta(days=int(v))
        except Exception:
            return None
    # 문자열인 경우 _parse_birthdate 사용
    return _parse_birthdate(str(v))

def _parse_birthdate(s: str | None) -> Optional[date]:
    if not s:
        return None
    s = str(s).strip().replace(" ", "")
    # 예: 2008.12.20. -> 2008-12-20
    s = s.strip(".").replace(".", "-")
    # 2008-12-20 / 2008-12 / 2008 등 다양한 경우가 있을 수 있어 보수적으로 처리
    fmts = ("%Y-%m-%d", "%Y-%m-%d.", "%Y-%m", "%Y.%m.%d", "%Y.%m")
    for fmt in fmts:
        try:
            dt = datetime.strptime(s, fmt)
            # 일 없는 형태면 1일로 보정
            day = dt.day if "%d" in fmt else 1
            month = dt.month if "%m" in fmt else 1
            return date(dt.year, month, day)
        except Exception:
            continue
    return None

def _parse_gender(s: str | None) -> Optional[str]:
    if not s:
        return None
    s = str(s).strip()
    if s.startswith("남"):
        return "M"
    if s.startswith("여"):
        return "F"
    if s.upper() in ("M", "F"):
        return s.upper()
    return None

def _must_be_homeroom_of(current: User, db: Session, grade: int, class_no: int) -> None:
    if current.role == "admin":
        return
    if current.role != "teacher":
        raise HTTPException(status_code=403, detail="교사만 업로드할 수 있습니다.")
    ha = (
        db.query(HomeroomAssignment)
        .filter(
            HomeroomAssignment.teacher_user_id == current.id,
            HomeroomAssignment.grade == grade,
            HomeroomAssignment.class_no == class_no,
        )
        .first()
    )
    if not ha:
        raise HTTPException(
            status_code=403,
            detail=f"담임 매핑이 없습니다: {grade}학년 {class_no}반 (관리자에게 문의)",
        )
    if not current.teacher_id:
        raise HTTPException(status_code=400, detail="교사 프로필(teacher_id)이 연결되지 않았습니다.")


# ------------------------- 템플릿 다운로드 -------------------------
@router.get("/homeroom/students/template.xlsx", dependencies=[Depends(role_required("teacher","admin"))])
def download_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "학생정보"
    ws.append(EXPECTED_HEADERS)
    ws.append([
        1, 9, 1, "홍길동", 2025000130, "여성", "2008.12.20.",
        "서울특별시 중구 세종대로 110", "특수학생", "010-0000-0000",
        "010-1111-1111", "010-2222-2222", "s2025000130",
    ])
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    real_name = "학생업로드_템플릿.xlsx"
    fallback = "student_template.xlsx"
    disposition = f"attachment; filename={fallback}; filename*=UTF-8''{quote(real_name)}"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": disposition},
    )


# ------------------------- 업로드/업서트 -------------------------
@router.post("/homeroom/students/upload-xlsx", dependencies=[Depends(role_required("teacher","admin"))])
def upload_students_xlsx(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current: User = Depends(get_current_user),
):
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="xlsx 파일을 업로드하세요.")
    try:
        wb = load_workbook(file.file, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="엑셀을 읽을 수 없습니다.")

    ws = wb.active
    # 헤더 검증 (왼쪽에서 EXPECTED_HEADERS 길이만큼 확인)
    header = [
        str((ws.cell(row=1, column=i + 1).value or "")).strip()
        for i in range(len(EXPECTED_HEADERS))
    ]
    if header != EXPECTED_HEADERS:
        raise HTTPException(
            status_code=400,
            detail=f"헤더가 일치하지 않습니다. 기대 순서: {', '.join(EXPECTED_HEADERS)}",
        )

    created, updated, skipped = [], [], []
    skipped_reasons: dict[str, str] = {}

    # values_only=True 로 바로 값만 받기, max_col로 과잉컬럼 방지
    for r, row in enumerate(ws.iter_rows(min_row=2, max_col=len(EXPECTED_HEADERS), values_only=True), start=2):
        # 완전 빈 행이면 스킵
        if not any(cell not in (None, "", " ") for cell in row):
            continue

        key = f"ROW{r}"
        try:
            grade = _safe_int(row[0])
            class_no = _safe_int(row[1])
            number = _safe_int(row[2])
            name = _safe_str(row[3])
            sid_raw = row[4]

            # 필수값 점검
            if grade is None:
                raise ValueError("학년 누락/숫자 아님")
            if class_no is None:
                raise ValueError("반 누락/숫자 아님")
            if number is None:
                raise ValueError("번호 누락/숫자 아님")
            if not name:
                raise ValueError("성명 누락")
            student_id = _safe_int(sid_raw)
            if student_id is None:
                raise ValueError("학생개인번호 누락/숫자 아님")

            # 담임 권한 확인
            try:
                _must_be_homeroom_of(current, db, grade, class_no)
            except HTTPException as he:
                raise ValueError(he.detail)

            gender = _parse_gender(_safe_str(row[5]))
            birthdate = _excel_date_to_date(row[6])
            address = _safe_str(row[7])
            # 비고 = row[8] (현재 미사용)
            phone = _safe_str(row[9])
            parent1 = _safe_str(row[10])
            parent2 = _safe_str(row[11])
            # username = _safe_str(row[12])  # 참고용, 저장/검증 안 함

            stu = db.get(Student, student_id)
            if stu is None:
                stu = Student(
                    id=student_id,
                    student_no=str(student_id),
                    name=name,
                    grade=grade,
                    class_no=class_no,
                    number=number,
                    gender=gender or "",
                    phone=phone,
                    parent1_phone=parent1,
                    parent2_phone=parent2,
                    address=address,
                    homeroom_teacher_id=current.teacher_id,
                )
                if hasattr(Student, "birthdate"):
                    setattr(stu, "birthdate", birthdate)
                db.add(stu)
                created.append(str(student_id))
            else:
                stu.name = name
                stu.grade = grade
                stu.class_no = class_no
                stu.number = number
                if gender:
                    stu.gender = gender
                stu.phone = phone
                stu.parent1_phone = parent1
                stu.parent2_phone = parent2
                stu.address = address
                if hasattr(Student, "birthdate"):
                    setattr(stu, "birthdate", birthdate)
                stu.homeroom_teacher_id = current.teacher_id
                updated.append(str(student_id))

        except ValueError as ve:
            skipped.append(key)
            skipped_reasons[key] = str(ve)
        except Exception:
            skipped.append(key)
            skipped_reasons[key] = "예상치 못한 형식 오류"

    db.commit()
    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "skipped_reasons": skipped_reasons,
    }
